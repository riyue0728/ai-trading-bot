# -*- coding: utf-8 -*-
"""
自动盯盘记录模块
- 记录每次交易信号
- 定时检查持仓状态
- 复盘结果，统计准确率
"""

import json
import time
from datetime import datetime
from pathlib import Path
import logging

logger = logging.getLogger("ClawBot")

# 记录文件
SIGNALS_FILE = Path(__file__).parent.parent / "signals_history.json"

# 记录模板
signal_record_template = {
    "signal_id": "",          # 信号唯一ID
    "timestamp": "",           # 信号时间
    "ticker": "",             # 交易对
    "direction": "",          # "long" / "short"
    "entry_price": 0,         # 入场价
    "stop_loss": 0,           # 止损价
    "take_profit_1": 0,       # T1止盈
    "take_profit_2": 0,       # T2止盈
    "ai_decision": "",        # AI决策
    "status": "pending",      # pending/active/closed
    "close_reason": "",       # tp/sl/expired/manual
    "close_price": 0,         # 平仓价
    "pnl_pct": 0,             # 盈亏百分比
    "result_time": "",        # 平仓时间
    "notes": ""               # 备注
}

def init_signals_file():
    """初始化信号记录文件"""
    if not SIGNALS_FILE.exists():
        SIGNALS_FILE.write_text(json.dumps([], ensure_ascii=False, indent=2))
        logger.info(f"创建信号记录文件: {SIGNALS_FILE}")

def save_signal(signal_data: dict):
    """保存新信号"""
    init_signals_file()
    
    signals = json.loads(SIGNALS_FILE.read_text(encoding='utf-8'))
    
    # 生成信号ID
    signal_id = f"{signal_data.get('ticker', 'UNK')}_{int(time.time())}"
    signal_data["signal_id"] = signal_id
    signal_data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    signal_data["status"] = "pending"
    
    signals.append(signal_data)
    SIGNALS_FILE.write_text(json.dumps(signals, ensure_ascii=False, indent=2))
    
    logger.info(f"📝 记录信号: {signal_id} - {signal_data.get('ticker')} {signal_data.get('direction')} @ {signal_data.get('entry_price')}")
    return signal_id

def update_signal_status(signal_id: str, status: str, close_price: float = 0, reason: str = ""):
    """更新信号状态"""
    init_signals_file()
    
    signals = json.loads(SIGNALS_FILE.read_text(encoding='utf-8'))
    
    for s in signals:
        if s.get("signal_id") == signal_id:
            s["status"] = status
            s["close_price"] = close_price
            s["close_reason"] = reason
            s["result_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 计算盈亏
            if close_price > 0 and s.get("entry_price", 0) > 0:
                if s.get("direction") == "long":
                    s["pnl_pct"] = round((close_price - s["entry_price"]) / s["entry_price"] * 100, 2)
                else:
                    s["pnl_pct"] = round((s["entry_price"] - close_price) / s["entry_price"] * 100, 2)
            
            SIGNALS_FILE.write_text(json.dumps(signals, ensure_ascii=False, indent=2))
            logger.info(f"🔄 更新信号: {signal_id} -> {status} ({reason})")
            return True
    
    return False

def get_active_signals():
    """获取未平仓信号"""
    init_signals_file()
    
    signals = json.loads(SIGNALS_FILE.read_text(encoding='utf-8'))
    return [s for s in signals if s.get("status") == "active"]

def get_pending_signals():
    """获取待激活信号"""
    init_signals_file()
    
    signals = json.loads(SIGNALS_FILE.read_text(encoding='utf-8'))
    return [s for s in signals if s.get("status") == "pending"]

def mark_signal_active(signal_id: str, entry_price: float):
    """标记信号已入场"""
    return update_signal_status(signal_id, "active", entry_price, "entered")

def close_signal(signal_id: str, close_price: float, reason: str):
    """平仓信号"""
    return update_signal_status(signal_id, "closed", close_price, reason)

def get_statistics():
    """获取统计信息"""
    init_signals_file()
    
    signals = json.loads(SIGNALS_FILE.read_text(encoding='utf-8'))
    
    total = len(signals)
    closed = [s for s in signals if s.get("status") == "closed"]
    active = [s for s in signals if s.get("status") == "active"]
    pending = [s for s in signals if s.get("status") == "pending"]
    
    wins = [s for s in closed if s.get("pnl_pct", 0) > 0]
    losses = [s for s in closed if s.get("pnl_pct", 0) <= 0]
    
    win_rate = len(wins) / len(closed) * 100 if closed else 0
    avg_pnl = sum(s.get("pnl_pct", 0) for s in closed) / len(closed) if closed else 0
    
    return {
        "total": total,
        "closed": len(closed),
        "active": len(active),
        "pending": len(pending),
        "wins": len(wins),
        "losses": len(losses),
        "win_rate": round(win_rate, 2),
        "avg_pnl": round(avg_pnl, 2)
    }

def export_to_excel():
    """导出到 Excel（需要 openpyxl）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "交易记录"
        
        # 表头
        headers = ["信号ID", "时间", "交易对", "方向", "入场价", "止损", "T1", "T2", "AI决策", "状态", "平仓原因", "平仓价", "盈亏%", "平仓时间", "备注"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 读取信号
        init_signals_file()
        signals = json.loads(SIGNALS_FILE.read_text(encoding='utf-8'))
        
        # 写入数据
        for row, s in enumerate(signals, 2):
            ws.cell(row=row, column=1, value=s.get("signal_id", ""))
            ws.cell(row=row, column=2, value=s.get("timestamp", ""))
            ws.cell(row=row, column=3, value=s.get("ticker", ""))
            ws.cell(row=row, column=4, value=s.get("direction", ""))
            ws.cell(row=row, column=5, value=s.get("entry_price", 0))
            ws.cell(row=row, column=6, value=s.get("stop_loss", 0))
            ws.cell(row=row, column=7, value=s.get("take_profit_1", 0))
            ws.cell(row=row, column=8, value=s.get("take_profit_2", 0))
            ws.cell(row=row, column=9, value=s.get("ai_decision", ""))
            ws.cell(row=row, column=10, value=s.get("status", ""))
            ws.cell(row=row, column=11, value=s.get("close_reason", ""))
            ws.cell(row=row, column=12, value=s.get("close_price", 0))
            ws.cell(row=row, column=13, value=s.get("pnl_pct", 0))
            ws.cell(row=row, column=14, value=s.get("result_time", ""))
            ws.cell(row=row, column=15, value=s.get("notes", ""))
        
        # 保存
        excel_path = Path(__file__).parent.parent / "trading_records.xlsx"
        wb.save(str(excel_path))
        logger.info(f"导出Excel: {excel_path}")
        return str(excel_path)
        
    except ImportError:
        logger.warning("需要安装 openpyxl: pip install openpyxl")
        return None

# 测试
if __name__ == "__main__":
    print("=== 盯盘记录测试 ===")
    
    # 保存测试信号
    test_signal = {
        "ticker": "XAU_USDT",
        "direction": "long",
        "entry_price": 4975.5,
        "stop_loss": 4965,
        "take_profit_1": 4990,
        "take_profit_2": 5010,
        "ai_decision": "强烈买入"
    }
    
    sid = save_signal(test_signal)
    print(f"保存信号: {sid}")
    
    # 模拟平仓
    close_signal(sid, 4995.5, "tp")
    print(f"标记止盈")
    
    # 统计
    stats = get_statistics()
    print(f"\n统计: {stats}")
