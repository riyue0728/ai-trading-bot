# -*- coding: utf-8 -*-
"""
黄金价格监控模块
- 获取实时价格（Yahoo Finance）
- 监控止盈/止损
- 记录到Excel
"""

import time
import json
import requests
import yfinance as yf
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import logging

logger = logging.getLogger("ClawdBot")

# GoldAPI.io 配置（免费黄金价格 API，每月50次）
GOLD_API_KEY = "goldapi-bjhc1smldqlqws-io"
GOLD_API_URL = "https://api.gold-api.com/price/XAU"

# Yahoo Finance 黄金代码
GOLD_TICKER = "XAUUSD=X"

# Excel 文件路径
EXCEL_FILE = "/root/.openclaw/workspace/ai-trading-bot/trading_records.xlsx"

def get_gold_price():
    """获取黄金价格（优先 GoldAPI.io，备用 Yahoo Finance）"""
    # 优先：GoldAPI.io
    try:
        headers = {"Authorization": f"Bearer {GOLD_API_KEY}"}
        response = requests.get(GOLD_API_URL, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            price = data.get('price')
            if price:
                return float(price)
    except Exception as e:
        logger.warning(f"GoldAPI.io 获取失败: {e}")
    
    # 备用：Yahoo Finance
    try:
        gold = yf.Ticker(GOLD_TICKER)
        price = gold.history(period="1m")
        if not price.empty:
            return float(price['Close'].iloc[-1])
    except Exception as e:
        logger.warning(f"Yahoo Finance 获取失败: {e}")
    
    return None

def init_excel():
    """初始化Excel文件"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "交易记录"
        
        # 表头
        headers = [
            "时间", "品种", "方向", "开仓价", "止损价", "止盈价(T1)", "止盈价(T2)",
            "触发价格", "触发类型", "盈亏(pips)", "状态", "AI决策"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(EXCEL_FILE)
        logger.info(f"✅ 创建Excel文件: {EXCEL_FILE}")

def add_trade_record(trade_info):
    """添加交易记录"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        row = ws.max_row + 1
        
        for col, key in enumerate([
            "time", "ticker", "direction", "entry_price", "stop_loss", "take_profit1", "take_profit2",
            "trigger_price", "trigger_type", "pnl", "status", "ai_decision"
        ], 1):
            ws.cell(row=row, column=col, value=trade_info.get(key, ""))
        
        wb.save(EXCEL_FILE)
        logger.info(f"✅ 记录交易: {trade_info['direction']} @ {trade_info['entry_price']}")
        
    except Exception as e:
        logger.error(f"❌ 记录失败: {e}")

def get_price_with_retry(max_retries=3):
    """带重试的价格获取"""
    for i in range(max_retries):
        price = get_gold_price()
        if price:
            return price
        time.sleep(1)
    return None

def monitor_trade(entry_price, direction, stop_loss, take_profit1, take_profit2, ai_decision):
    """
    监控交易
    direction: "做多" 或 "做空"
    """
    logger.info(f"\n🔔 开始监控交易:")
    logger.info(f"   方向: {direction}")
    logger.info(f"   开仓价: {entry_price}")
    logger.info(f"   止损: {stop_loss}")
    logger.info(f"   止盈T1: {take_profit1}")
    logger.info(f"   止盈T2: {take_profit2}")
    
    # 记录交易
    trade_info = {
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ticker": "XAUUSD",
        "direction": direction,
        "entry_price": entry_price,
        "stop_loss": stop_loss,
        "take_profit1": take_profit1,
        "take_profit2": take_profit2,
        "trigger_price": "",
        "trigger_type": "",
        "pnl": "",
        "status": "监控中",
        "ai_decision": ai_decision
    }
    add_trade_record(trade_info)
    
    # 等待价格
    current_price = get_price_with_retry()
    if not current_price:
        logger.error("❌ 无法获取价格，监控失败")
        return
    
    logger.info(f"   当前价格: {current_price}")
    
    return current_price

# 测试
if __name__ == "__main__":
    init_excel()
    price = get_gold_price()
    if price:
        print(f"✅ 黄金价格: ${price}")
    else:
        print("❌ 获取价格失败")

